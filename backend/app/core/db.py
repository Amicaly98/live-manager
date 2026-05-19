"""
db.py - SQLite 数据库层

替代 Excel 作为任务数据的主存储。
提供完整的 CRUD、迁移、导入导出能力。
"""

import sqlite3
import logging
import threading
from pathlib import Path
from typing import Optional, List, Dict, Any
from datetime import date, datetime
from contextlib import contextmanager

logger = logging.getLogger(__name__)

DEFAULT_DB_PATH = Path("live_tasks.db")

# ==================== Schema ====================

CREATE_TASKS_TABLE = """
CREATE TABLE IF NOT EXISTS tasks (
id INTEGER PRIMARY KEY AUTOINCREMENT,
zone_name TEXT NOT NULL UNIQUE,
category INTEGER NOT NULL DEFAULT 1,
total_days INTEGER NOT NULL DEFAULT 1,
days_done INTEGER NOT NULL DEFAULT 0,
deadline_raw TEXT NOT NULL DEFAULT '',
today_done INTEGER DEFAULT NULL,
remaining_days INTEGER NOT NULL DEFAULT 1,
-- 计算列（排序用，由 _compute_static 更新）
a_val INTEGER NOT NULL DEFAULT 9999,
i_val INTEGER NOT NULL DEFAULT 0,
j_val INTEGER NOT NULL DEFAULT -1,
created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);
"""

CREATE_INDEX_ZONE = """
CREATE INDEX IF NOT EXISTS idx_tasks_zone ON tasks(zone_name);
"""

CREATE_INDEX_PRIORITY = """
CREATE INDEX IF NOT EXISTS idx_tasks_priority ON tasks(a_val, category);
"""


class TaskDB:
    """SQLite 任务数据库管理器（线程安全）"""

    def __init__(self, db_path: str = None):
        self.db_path = Path(db_path or DEFAULT_DB_PATH).resolve()
        self._lock = threading.Lock()
        self._init_db()

    # ==================== 初始化 ====================

    def _init_db(self):
        """创建表和索引，自动迁移旧 schema"""
        with self._get_conn() as conn:
            cols = [r[1] for r in conn.execute(
                "PRAGMA table_info(tasks)").fetchall()]
            if 'extra_hours' in cols or 'priority' in cols:
                logger.info("检测到旧版 schema，正在迁移...")
                old_data = []
                if cols:
                    old_data = [dict(r) for r in conn.execute(
                        "SELECT * FROM tasks").fetchall()]
                conn.execute("DROP TABLE IF EXISTS tasks")
                conn.execute(CREATE_TASKS_TABLE)
                for row in old_data:
                    conn.execute(
                        """INSERT INTO tasks
                        (zone_name, category, total_days, days_done,
                        deadline_raw, today_done, remaining_days,
                        a_val, i_val, j_val)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (
                            row.get('zone_name', ''),
                            row.get('category', 1),
                            row.get('total_days', 1),
                            row.get('days_done', 0),
                            row.get('deadline_raw', ''),
                            row.get('today_done'),
                            row.get('remaining_days', 1),
                            row.get('a_val', 9999),
                            row.get('i_val', 0),
                            row.get('j_val', -1),
                        )
                    )
                conn.commit()
                logger.info(f"迁移完成，已恢复 {len(old_data)} 条任务")
            else:
                conn.execute(CREATE_TASKS_TABLE)
                conn.execute(CREATE_INDEX_ZONE)
                conn.execute(CREATE_INDEX_PRIORITY)
                conn.commit()
                logger.info(f"数据库就绪：{self.db_path}")

    @contextmanager
    def _get_conn(self):
        """获取数据库连接（线程安全）"""
        conn = sqlite3.connect(str(self.db_path))
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            yield conn
        finally:
            conn.close()

    # ==================== 基础 CRUD ====================

    def get_all_tasks(self) -> List[Dict[str, Any]]:
        with self._get_conn() as conn:
            rows = conn.execute(
                "SELECT * FROM tasks ORDER BY a_val ASC"
            ).fetchall()
            return [dict(r) for r in rows]

    def get_task_by_zone(self, zone_name: str) -> Optional[Dict[str, Any]]:
        with self._get_conn() as conn:
            row = conn.execute(
                "SELECT * FROM tasks WHERE zone_name = ?", (zone_name,)
            ).fetchone()
            return dict(row) if row else None

    def get_task_by_id(self, task_id: int) -> Optional[Dict[str, Any]]:
        with self._get_conn() as conn:
            row = conn.execute(
                "SELECT * FROM tasks WHERE id = ?", (task_id,)
            ).fetchone()
            return dict(row) if row else None

    def insert_task(self, task_data: dict) -> int:
        with self._get_conn() as conn:
            cursor = conn.execute(
                """INSERT INTO tasks
                (zone_name, category, total_days, days_done, deadline_raw,
                today_done, remaining_days)
                VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    task_data['zone_name'],
                    task_data.get('category', 1),
                    task_data.get('total_days', 1),
                    task_data.get('days_done', 0),
                    task_data.get('deadline_raw', ''),
                    task_data.get('today_done'),
                    task_data.get('remaining_days', 1),
                )
            )
            conn.commit()
            return cursor.lastrowid

    def update_task(self, zone_name: str, updates: dict) -> bool:
        allowed = {
            'category', 'total_days', 'days_done', 'deadline_raw',
            'today_done', 'remaining_days',
            'a_val', 'i_val', 'j_val'
        }
        filtered = {k: v for k, v in updates.items() if k in allowed}
        if not filtered:
            return False
        filtered['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        set_clause = ', '.join(f"{k} = ?" for k in filtered)
        values = list(filtered.values()) + [zone_name]
        with self._get_conn() as conn:
            cursor = conn.execute(
                f"UPDATE tasks SET {set_clause} WHERE zone_name = ?", values
            )
            conn.commit()
            return cursor.rowcount > 0

    def update_task_by_id(self, task_id: int, updates: dict) -> bool:
        allowed = {
            'zone_name', 'category', 'total_days', 'days_done', 'deadline_raw',
            'today_done', 'remaining_days',
            'a_val', 'i_val', 'j_val'
        }
        filtered = {k: v for k, v in updates.items() if k in allowed}
        if not filtered:
            return False
        filtered['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        set_clause = ', '.join(f"{k} = ?" for k in filtered)
        values = list(filtered.values()) + [task_id]
        with self._get_conn() as conn:
            cursor = conn.execute(
                f"UPDATE tasks SET {set_clause} WHERE id = ?", values
            )
            conn.commit()
            return cursor.rowcount > 0

    def delete_task(self, zone_name: str) -> bool:
        with self._get_conn() as conn:
            cursor = conn.execute(
                "DELETE FROM tasks WHERE zone_name = ?", (zone_name,)
            )
            conn.commit()
            return cursor.rowcount > 0

    def delete_task_by_id(self, task_id: int) -> bool:
        with self._get_conn() as conn:
            cursor = conn.execute(
                "DELETE FROM tasks WHERE id = ?", (task_id,)
            )
            conn.commit()
            return cursor.rowcount > 0

    def clear_today_done(self):
        with self._get_conn() as conn:
            conn.execute(
                "UPDATE tasks SET today_done = NULL, updated_at = ?",
                (datetime.now().strftime('%Y-%m-%d %H:%M:%S'),)
            )
            conn.commit()
            logger.info("已清空 today_done 标志")

    def update_static_values(self, computed: List[dict]):
        with self._get_conn() as conn:
            for item in computed:
                conn.execute(
                    """UPDATE tasks
                    SET a_val = ?, i_val = ?, j_val = ?,
                    updated_at = ?
                    WHERE zone_name = ?""",
                    (
                        item['a'], item['i'], item['j'],
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        item['zone_name'],
                    )
                )
            conn.commit()
            logger.info(f"静态值已更新（{len(computed)} 行）")

    def delete_all_tasks(self):
        with self._get_conn() as conn:
            conn.execute("DELETE FROM tasks")
            conn.commit()
            logger.info("已清空所有任务")

    def count_tasks(self) -> int:
        with self._get_conn() as conn:
            return conn.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]

    def has_tasks(self) -> bool:
        return self.count_tasks() > 0

    @staticmethod
    def read_zone_names_from_excel(excel_path: str) -> List[str]:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        zones = []
        for row in range(2, ws.max_row + 1):
            zone = str(ws.cell(row=row, column=2).value or '').strip()
            if zone:
                zones.append(zone)
        wb.close()
        return zones

    def import_from_excel(self, excel_path: str, skip_zones: List[str] = None) -> dict:
        """从 Excel 导入任务到数据库，返回 {imported, updated, skipped, errors}
        校验规则：
        - B列（分区名）为空 跳过
        - D列（需要完成天数）为空或0 跳过并记录错误
        - F列（截止时间）无法解析为有效日期 跳过并记录错误
        - C列（类别）默认2（非数字或<=0 默认2）
        - skip_zones 中的分区名 跳过
        """
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        skip_set = set(skip_zones or [])

        result = {'imported': 0, 'updated': 0, 'skipped': 0, 'errors': []}
        with self._get_conn() as conn:
            for row in range(2, ws.max_row + 1):
                zone_name = str(ws.cell(row=row, column=2).value or '').strip()
                if not zone_name:
                    continue

                if zone_name in skip_set:
                    continue

                # C列：类别，默认2（仅空/非数字时默认2，0=已完成是合法值）
                category = self._safe_int(ws.cell(row=row, column=3).value, None)
                if category is None:
                    category = 2

                # D列：需要完成天数，不能为空
                total_days = self._safe_int(ws.cell(row=row, column=4).value, None)
                if total_days is None or total_days <= 0:
                    result['skipped'] += 1
                    result['errors'].append(
                        f"第{row}行「{zone_name}」：需要完成天数无效，已跳过")
                    continue

                days_done = self._safe_int(ws.cell(row=row, column=5).value, 0)

                # F列：截止时间，必须能解析为有效日期
                deadline_raw = str(ws.cell(row=row, column=6).value or '').strip()
                if not deadline_raw or not self._validate_deadline(deadline_raw):
                    result['skipped'] += 1
                    result['errors'].append(
                        f"第{row}行「{zone_name}」：截止时间无效（'{deadline_raw}'），已跳过")
                    continue

                today_done = self._safe_int_or_none(ws.cell(row=row, column=7).value)

                existing = conn.execute(
                    "SELECT id FROM tasks WHERE zone_name = ?", (zone_name,)
                ).fetchone()

                if existing:
                    conn.execute(
                        """UPDATE tasks SET
                        category=?, total_days=?, days_done=?,
                        deadline_raw=?, today_done=?,
                        updated_at=?
                        WHERE zone_name=?""",
                        (category, total_days, days_done,
                         deadline_raw, today_done,
                         datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                         zone_name)
                    )
                    result['updated'] += 1
                else:
                    conn.execute(
                        """INSERT INTO tasks
                        (zone_name, category, total_days, days_done,
                        deadline_raw, today_done)
                        VALUES (?, ?, ?, ?, ?, ?)""",
                        (zone_name, category, total_days, days_done,
                         deadline_raw, today_done)
                    )
                    result['imported'] += 1

            conn.commit()
        wb.close()

        logger.info(
            f"从 Excel 导入：新增 {result['imported']}，"
            f"更新 {result['updated']}，跳过 {result['skipped']}")
        if result['errors']:
            for err in result['errors']:
                logger.warning(f" {err}")
        return result

    @staticmethod
    def _validate_deadline(raw_val: str) -> bool:
        import re as _re
        from datetime import date as _date, datetime as _dt, timedelta as _td
        s = raw_val.strip()
        if not s:
            return False
        m = _re.match(r'=DATE\((\d+),(\d+),(\d+)\)', s, _re.IGNORECASE)
        if m:
            try:
                _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                return True
            except ValueError:
                return False
        try:
            serial = int(float(s))
            if 40000 < serial < 80000:
                return True
        except BaseException:
            pass
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
            try:
                _dt.strptime(s, fmt)
                return True
            except ValueError:
                pass
        return False

    def export_to_excel(self, excel_path: str, stats: dict = None):
        """导出任务到 Excel
        格式：宋体 16pt；任务行 A-I + J/K 列放统计值
        列宽 = 表头除括号外字数（分区名=10汉字）
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        font = Font(name='宋体', size=16)
        font_bold = Font(name='宋体', size=16, bold=True)
        align_left = Alignment(horizontal='left', vertical='center')
        align_center = Alignment(horizontal='center', vertical='center')

        # 表头
        headers = [
            '优先度',  # A
            '分区名',  # B（居中）
            '类别(每天需要完成小时数，默认2，0为已完成状态)',  # C
            '需要完成天数',  # D
            '已完成天数',  # E
            '截止时间（默认为当天的23：59）',  # F
            '今日是否完成（完成扣1）',  # G
            '距离完成',  # H
            '松弛度',  # I
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_bold
            if col_idx == 2:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        # 数据行
        tasks = self.get_all_tasks()
        for i, t in enumerate(tasks, start=2):
            j_val = t.get('j_val', -1)
            ws.cell(row=i, column=1, value=t.get('a_val', 9999)).font = font
            ws.cell(row=i, column=2, value=t['zone_name']).font = font
            ws.cell(row=i, column=3, value=t['category']).font = font
            ws.cell(row=i, column=4, value=t['total_days']).font = font
            ws.cell(row=i, column=5, value=t['days_done']).font = font
            dl_str = self._format_deadline(t.get('deadline_raw', ''))
            ws.cell(row=i, column=6, value=dl_str).font = font
            ws.cell(row=i, column=7, value=t.get('today_done')).font = font
            ws.cell(row=i, column=8, value=t.get('i_val', 0)).font = font
            ws.cell(row=i, column=9, value=j_val).font = font
            for col in range(1, 10):
                ws.cell(row=i, column=col).alignment = align_left

        # 统计值：J/K 列，第 1-3 行显示
        stats = stats or {}
        stat_font = Font(name='宋体', size=16, bold=True, color='CC0000')

        stat_items = [
            ('剩余时间：', stats.get('remaining_time', 0)),
            ('平均剩余时间：', round(stats.get('avg_remaining', 0), 2)),
            ('紧迫率：', f"{round(stats.get('urgency', 0) * 100, 2)}%"),
        ]

        for idx, (item_label, item_value) in enumerate(stat_items):
            row_num = idx + 1
            cell_j = ws.cell(row=row_num, column=10, value=item_label)
            cell_j.font = stat_font
            cell_j.alignment = align_center
            cell_k = ws.cell(row=row_num, column=11, value=item_value)
            cell_k.font = stat_font
            cell_k.alignment = align_center

        # 列宽
        widths = {
            'A': 10, 'B': 30, 'C': 7,
            'D': int(6 * 3.3), 'E': int(5 * 3.3),
            'F': 15, 'G': int(6 * 3.3),
            'H': int(4 * 3.3), 'I': 10,
            'J': int(7 * 3.3), 'K': int(4 * 3.3),
        }
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        wb.save(excel_path)
        wb.close()
        logger.info(f"导出 {len(tasks)} 条任务到 {excel_path}")

    @staticmethod
    def _format_deadline(raw_val: str) -> str:
        import re as _re
        from datetime import date as _date, datetime as _dt, timedelta as _td

        s = raw_val.strip() if raw_val else ''
        if not s:
            return ''
        # =DATE(Y,M,D)
        m = _re.match(r'=DATE\((\d+),(\d+),(\d+)\)', s, _re.IGNORECASE)
        if m:
            try:
                d = _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                return f"{d.year}/{d.month}/{d.day}"
            except ValueError:
                return s
        # 序列号
        try:
            serial = int(float(s))
            if 40000 < serial < 80000:
                d = _date(1899, 12, 30) + _td(days=serial)
                return f"{d.year}/{d.month}/{d.day}"
        except BaseException:
            pass
        # 日期字符串（含时间部分）
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y/%m/%d']:
            try:
                d = _dt.strptime(s, fmt).date()
                return f"{d.year}/{d.month}/{d.day}"
            except ValueError:
                pass
        return s

    def auto_migrate_from_excel(self, excel_path: str) -> bool:
        if self.has_tasks():
            logger.info("数据库已有数据，跳过迁移")
            return False
        excel = Path(excel_path)
        if not excel.exists():
            logger.info("Excel 文件不存在，跳过迁移")
            return False
        logger.info(f"检测到 Excel 文件，开始自动迁移：{excel}")
        result = self.import_from_excel(str(excel))
        logger.info(
            f"自动迁移完成：新增 {result['imported']}，"
            f"更新 {result['updated']}，跳过 {result['skipped']}")
        return result['imported'] + result['updated'] > 0

    @staticmethod
    def _safe_int(value, default=0):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _safe_int_or_none(value):
        if value is None:
            return None
        try:
            return int(float(value))
        except (TypeError, ValueError):
            s = str(value).strip()
            if s == '1':
                return 1
            return None
